#!/usr/bin/env python3
"""
Flask Backend for Barcode Generator Web UI
Connects the HTML frontend with barcode generation functions
"""

from flask import Flask, render_template, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image, ImageDraw
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
import os
import tempfile
import shutil
from io import BytesIO
from datetime import datetime
import mimetypes

app = Flask(__name__, static_folder='.')
CORS(app)

# Configuration
OUTPUT_FOLDER = 'outputs'

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def generate_barcode_image(data, width=200, height=100):
    """Generate a barcode image from data string"""
    try:
        from barcode import Code128
        from barcode.writer import ImageWriter
        
        barcode_instance = Code128(data, writer=ImageWriter())
        image_bytes = BytesIO()
        barcode_instance.write(image_bytes)
        image_bytes.seek(0)
        
        img = Image.open(image_bytes)
        img = img.convert('RGB')
        return img
        
    except ImportError:
        # Fallback: Create simple barcode pattern
        img = Image.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(img)
        
        bar_width = width // len(data)
        bar_height = int(height * 0.7)
        
        for i, char in enumerate(data):
            ascii_val = ord(char)
            bar_count = ascii_val % 8
            x_start = i * bar_width
            
            for j in range(bar_count):
                bar_x = x_start + (j * (bar_width // bar_count))
                draw.rectangle([bar_x, 10, bar_x + 2, bar_x + bar_height], fill='black')
        
        draw.text((10, height - 20), data, fill='black')
        return img

def create_excel_with_barcodes(numbers, output_path):
    """Create Excel file with barcodes"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Barcodes"
    
    # Headers
    ws['A1'] = "Number"
    ws['B1'] = "Barcode"
    ws['C1'] = "Barcode Code"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col in ['A1', 'B1', 'C1']:
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.row_dimensions[1].height = 25
    
    temp_dir = tempfile.mkdtemp(prefix="barcode_")
    
    try:
        for idx, number in enumerate(numbers, start=2):
            ws[f'A{idx}'] = number
            
            try:
                barcode_img = generate_barcode_image(str(number), width=250, height=100)
                barcode_filename = os.path.join(temp_dir, f"barcode_{number}.png")
                barcode_img.save(barcode_filename, 'PNG')
                
                if os.path.exists(barcode_filename) and os.path.getsize(barcode_filename) > 0:
                    img = XLImage(barcode_filename)
                    img.width = 250
                    img.height = 80
                    ws.add_image(img, f'B{idx}')
                    
                    ws[f'C{idx}'] = f"Code128: {number}"
                    ws[f'C{idx}'].alignment = Alignment(horizontal="left", vertical="center")
                    ws.row_dimensions[idx].height = 90
                    
            except Exception as e:
                ws[f'B{idx}'] = f"Error"
                ws[f'C{idx}'] = f"Error: {str(e)[:30]}"
        
        wb.save(output_path)
        return True
        
    finally:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)

def generate_pdf_from_numbers(numbers, cols=4, rows=8):
    """Generate A4 PDF with barcode layout"""
    temp_dir = tempfile.mkdtemp(prefix="barcode_pdf_")
    
    try:
        barcode_images = {}
        for number in numbers:
            try:
                barcode_img = generate_barcode_image(str(number), width=180, height=80)
                barcode_path = os.path.join(temp_dir, f"barcode_{number}.png")
                barcode_img.save(barcode_path, 'PNG')
                barcode_images[number] = barcode_path
            except:
                pass
        
        pdf_bytes = BytesIO()
        c = canvas.Canvas(pdf_bytes, pagesize=A4)
        page_width, page_height = A4
        
        margin_x = 12 * mm
        margin_y = 15 * mm
        
        available_width = page_width - (2 * margin_x)
        available_height = page_height - (2 * margin_y)
        
        cell_width = available_width / cols
        cell_height = available_height / rows
        
        for idx, number in enumerate(numbers):
            if number not in barcode_images:
                continue
            
            row = (idx % (cols * rows)) // cols
            col = (idx % (cols * rows)) % cols
            
            if idx > 0 and idx % (cols * rows) == 0:
                c.showPage()
                row = 0
                col = 0
            
            x = margin_x + (col * cell_width) + (cell_width * 0.05)
            y = page_height - margin_y - ((row + 1) * cell_height) + (cell_height * 0.05)
            
            barcode_width = cell_width * 0.9
            barcode_height = cell_height * 0.75
            
            barcode_path = barcode_images[number]
            try:
                c.drawImage(barcode_path, x, y, width=barcode_width, height=barcode_height, preserveAspectRatio=True)
                
                text_y = y - (cell_height * 0.15)
                c.setFont("Helvetica", 8)
                c.drawString(x + (barcode_width / 2) - 15, text_y, str(number))
            except:
                pass
        
        c.save()
        pdf_bytes.seek(0)
        return pdf_bytes
        
    finally:
        if os.path.exists(temp_dir):
            shutil.rmtree(temp_dir)

# ===================== API ROUTES =====================

@app.route('/')
def index():
    """Serve the main HTML page"""
    return send_file('barcode_generator_ui.html')

@app.route('/api/create-excel', methods=['POST'])
def create_excel_api():
    """Create a new Excel file with barcodes"""
    try:
        data = request.get_json()
        file_name = data.get('fileName', 'barcodes.xlsx')
        start_num = int(data.get('startNum', 253310001))
        count = int(data.get('count', 10))
        
        if count < 1 or count > 5000:
            return jsonify({'status': 'error', 'message': 'Count must be between 1 and 5000'}), 400
        
        numbers = [start_num + i for i in range(count)]
        
        output_path = os.path.join(OUTPUT_FOLDER, file_name)
        create_excel_with_barcodes(numbers, output_path)
        
        return jsonify({
            'status': 'success',
            'message': f'Excel file created with {count} barcodes',
            'fileName': file_name,
            'fileUrl': f'/download/{file_name}',
            'numbersGenerated': count,
            'range': f'{start_num} - {start_num + count - 1}'
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/upload-excel', methods=['POST'])
def upload_excel_api():
    """Upload Excel file to outputs folder and analyze it"""
    try:
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file provided'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No file selected'}), 400
        
        filename = file.filename
        # Save to outputs folder (not uploads)
        filepath = os.path.join(OUTPUT_FOLDER, filename)
        file.save(filepath)
        
        print(f"📁 File uploaded to: {filepath}")
        
        # Analyze file
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        numbers = []
        for row in range(2, ws.max_row + 1):
            cell_value = ws[f'A{row}'].value
            if cell_value is not None:
                try:
                    numbers.append(int(cell_value))
                except:
                    pass
        
        if not numbers:
            return jsonify({'status': 'error', 'message': 'No numbers found in Excel file'}), 400
        
        numbers.sort()
        last_number = max(numbers)
        
        return jsonify({
            'status': 'success',
            'message': 'File uploaded and analyzed',
            'fileName': filename,
            'filePath': filepath,
            'numbersFound': len(numbers),
            'range': f'{min(numbers)} - {max(numbers)}',
            'lastNumber': last_number,
            'numbers': numbers
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/update-excel', methods=['POST'])
def update_excel_api():
    """Update the uploaded Excel file with new numbers and generate barcodes"""
    try:
        data = request.get_json()
        file_path = data.get('filePath')
        new_start_num = int(data.get('newStartNum', 253310011))
        count = int(data.get('count', 10))
        
        if not os.path.exists(file_path):
            return jsonify({'status': 'error', 'message': f'File not found at: {file_path}'}), 404
        
        print(f"🔄 Updating file: {file_path}")
        
        # Load the uploaded file
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Find the next empty row (after existing data)
        next_row = ws.max_row + 1
        
        # Generate new barcodes
        temp_dir = tempfile.mkdtemp(prefix="barcode_")
        barcodes_added = 0
        
        try:
            for i in range(count):
                number = new_start_num + i
                row = next_row + i
                
                # Add the number to column A
                ws[f'A{row}'] = number
                
                try:
                    # Generate barcode image
                    barcode_img = generate_barcode_image(str(number), width=250, height=100)
                    barcode_filename = os.path.join(temp_dir, f"barcode_{number}.png")
                    barcode_img.save(barcode_filename, 'PNG')
                    
                    if os.path.exists(barcode_filename) and os.path.getsize(barcode_filename) > 0:
                        # Add barcode image to column B
                        img = XLImage(barcode_filename)
                        img.width = 250
                        img.height = 80
                        ws.add_image(img, f'B{row}')
                        
                        # Add barcode code to column C
                        ws[f'C{row}'] = f"Code128: {number}"
                        ws[f'C{row}'].alignment = Alignment(horizontal="left", vertical="center")
                        ws.row_dimensions[row].height = 90
                        
                        barcodes_added += 1
                        
                except Exception as e:
                    print(f"Error generating barcode for {number}: {e}")
                    ws[f'B{row}'] = "Error"
                    ws[f'C{row}'] = f"Error: {str(e)[:30]}"
            
            # SAVE THE FILE IN PLACE
            wb.save(file_path)
            print(f"✅ File saved: {file_path}")
            print(f"✅ Barcodes added: {barcodes_added}")
            
            # Return the filename for download (user will rename if needed)
            filename = os.path.basename(file_path)
            
            return jsonify({
                'status': 'success',
                'message': f'✅ Added {barcodes_added} new barcode(s)',
                'barcodesAdded': barcodes_added,
                'range': f'{new_start_num} - {new_start_num + count - 1}',
                'fileName': filename,
                'filePath': file_path
            })
            
        finally:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
        
    except Exception as e:
        print(f"❌ Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/rename-and-download', methods=['POST'])
def rename_and_download_api():
    """Rename file and prepare for download"""
    try:
        data = request.get_json()
        old_file_path = data.get('filePath')
        new_filename = data.get('newFileName')
        
        if not old_file_path or not new_filename:
            return jsonify({'status': 'error', 'message': 'Missing file path or new filename'}), 400
        
        if not os.path.exists(old_file_path):
            return jsonify({'status': 'error', 'message': 'File not found'}), 404
        
        # Ensure new filename has .xlsx extension
        if not new_filename.endswith(('.xlsx', '.xls')):
            new_filename += '.xlsx'
        
        new_file_path = os.path.join(OUTPUT_FOLDER, new_filename)
        
        # Check if new filename already exists
        if os.path.exists(new_file_path) and new_file_path != old_file_path:
            return jsonify({'status': 'error', 'message': f'File "{new_filename}" already exists'}), 400
        
        # If old and new paths are different, rename the file
        if old_file_path != new_file_path:
            os.rename(old_file_path, new_file_path)
            print(f"✅ File renamed: {os.path.basename(old_file_path)} → {new_filename}")
        
        return jsonify({
            'status': 'success',
            'message': f'File ready to download',
            'fileName': new_filename,
            'fileUrl': f'/download/{new_filename}'
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/generate-pdf', methods=['POST'])
def generate_pdf_api():
    """Generate A4 PDF with barcode layout"""
    try:
        data = request.get_json()
        excel_source = data.get('excelSource')
        pdf_name = data.get('pdfName', 'barcodes_a4.pdf')
        cols = int(data.get('cols', 4))
        rows = int(data.get('rows', 8))
        
        # Find the Excel file
        excel_path = os.path.join(OUTPUT_FOLDER, os.path.basename(excel_source))
        
        if not os.path.exists(excel_path):
            return jsonify({'status': 'error', 'message': f'Excel file not found: {excel_source}'}), 404
        
        # Read numbers from Excel
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        numbers = []
        for row in range(2, ws.max_row + 1):
            cell_value = ws[f'A{row}'].value
            if cell_value is not None:
                try:
                    numbers.append(str(int(cell_value)))
                except:
                    pass
        
        if not numbers:
            return jsonify({'status': 'error', 'message': 'No numbers found in Excel file'}), 400
        
        # Generate PDF
        pdf_bytes = generate_pdf_from_numbers(numbers, cols, rows)
        
        # Save to outputs folder
        output_path = os.path.join(OUTPUT_FOLDER, pdf_name)
        with open(output_path, 'wb') as f:
            f.write(pdf_bytes.getvalue())
        
        return jsonify({
            'status': 'success',
            'message': 'PDF generated successfully',
            'fileName': pdf_name,
            'fileUrl': f'/download/{pdf_name}',
            'barcodesInPDF': len(numbers),
            'layout': f'{cols}x{rows}'
        })
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    """Download generated or updated file"""
    try:
        file_path = os.path.join(OUTPUT_FOLDER, filename)
        
        if not os.path.exists(file_path):
            return jsonify({'status': 'error', 'message': 'File not found'}), 404
        
        print(f"📥 Downloading: {file_path}")
        
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/list-excel-files', methods=['GET'])
def list_excel_files():
    """List all available Excel files in outputs folder"""
    try:
        files = []
        
        # From outputs folder only (since all files are there now)
        if os.path.exists(OUTPUT_FOLDER):
            for f in os.listdir(OUTPUT_FOLDER):
                if f.endswith(('.xlsx', '.xls')):
                    files.append({'name': f})
        
        return jsonify({
            'status': 'success',
            'files': files
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({'status': 'healthy', 'message': 'Barcode Generator API is running'})

if __name__ == '__main__':
    print("\n" + "="*60)
    print("🚀 Barcode Generator - Web Interface")
    print("="*60)
    print("\n📍 Server starting on http://127.0.0.1:5000")
    print("\n✨ Features:")
    print("   ✅ Create new Excel files with barcodes")
    print("   ✅ Upload and UPDATE existing Excel files")
    print("   ✅ Auto-fill next number based on uploaded file")
    print("   ✅ Rename files before downloading")
    print("   ✅ Generate professional A4 PDFs")
    print("   ✅ Beautiful web interface")
    print("\n📂 Folders created:")
    print(f"   ✅ outputs/ - for ALL files (created & uploaded)")
    print("\n💡 Open your browser and go to:")
    print("   http://127.0.0.1:5000")
    print("\n⏹️  Press Ctrl+C to stop the server")
    print("="*60 + "\n")
    
    app.run(debug=True, host='127.0.0.1', port=5000)