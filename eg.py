#!/usr/bin/env python3
"""
Barcode Generator for Excel & PDF - A4 Layout (4x8)
Generates Code128 barcodes and creates printable A4 PDF with 4x8 layout
Usage: python generate_barcodes.py
"""


import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment, PatternFill
from PIL import Image, ImageDraw
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import inch, mm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
import os
import tempfile
import shutil


def encode_code128(data):
    """Encode data string to Code128 barcode format"""
    # Code128 character set and encoding
    code128_table = {
        ' ': 0, '!': 1, '"': 2, '#': 3, '$': 4, '%': 5, '&': 6, "'": 7, '(': 8, ')': 9,
        '*': 10, '+': 11, ',': 12, '-': 13, '.': 14, '/': 15, '0': 16, '1': 17, '2': 18, '3': 19,
        '4': 20, '5': 21, '6': 22, '7': 23, '8': 24, '9': 25, ':': 26, ';': 27, '<': 28, '=': 29,
        '>': 30, '?': 31, '@': 32, 'A': 33, 'B': 34, 'C': 35, 'D': 36, 'E': 37, 'F': 38, 'G': 39,
        'H': 40, 'I': 41, 'J': 42, 'K': 43, 'L': 44, 'M': 45, 'N': 46, 'O': 47, 'P': 48, 'Q': 49,
        'R': 50, 'S': 51, 'T': 52, 'U': 53, 'V': 54, 'W': 55, 'X': 56, 'Y': 57, 'Z': 58, '[': 59,
        '\\': 60, ']': 61, '^': 62, '_': 63
    }
    
    # Encode the data
    encoded = []
    checksum = 104  # Start Code B
    
    for i, char in enumerate(data):
        code = code128_table.get(char, 0)
        encoded.append(code)
        if i > 0:
            checksum += code * (i + 1)
    
    # Add checksum
    checksum = checksum % 103
    encoded.append(checksum)
    encoded.append(106)  # Stop code
    
    return encoded


def generate_barcode_image(data, width=200, height=100):
    """Generate a barcode image from data string"""
    try:
        # Use python-barcode library if available
        from barcode import Code128
        from barcode.writer import ImageWriter
        from io import BytesIO
        
        # Generate barcode
        barcode_instance = Code128(data, writer=ImageWriter())
        
        # Render to image
        image_bytes = BytesIO()
        barcode_instance.write(image_bytes)
        image_bytes.seek(0)
        
        img = Image.open(image_bytes)
        img = img.convert('RGB')
        
        return img
        
    except ImportError:
        # Fallback: Create simple barcode pattern with PIL
        img = Image.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Create simple visual barcode pattern
        bar_width = width // len(data)
        bar_height = height * 0.7
        
        for i, char in enumerate(data):
            ascii_val = ord(char)
            bar_count = ascii_val % 8
            
            x_start = i * bar_width
            
            for j in range(bar_count):
                bar_x = x_start + (j * (bar_width // bar_count))
                draw.rectangle([bar_x, 10, bar_x + 2, bar_x + bar_height], fill='black')
        
        # Add text below barcode
        draw.text((10, height - 20), data, fill='black')
        
        return img


def generate_pdf_from_excel(excel_file, pdf_file="barcodes_a4.pdf", cols=4, rows=8):
    """Generate A4 PDF with 4x8 barcode layout from Excel file"""
    
    print(f"📄 Generating PDF from {excel_file}...")
    
    # Load Excel file
    if not os.path.exists(excel_file):
        print(f"❌ Excel file not found: {excel_file}")
        return
    
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # Extract all numbers from column A
    numbers = []
    for row in range(2, ws.max_row + 1):
        cell_value = ws[f'A{row}'].value
        if cell_value is not None:
            try:
                numbers.append(str(int(cell_value)))
            except (ValueError, TypeError):
                pass
    
    if not numbers:
        print("❌ No numbers found in Excel file")
        return
    
    print(f"📊 Found {len(numbers)} barcodes to print")
    
    # Create temporary directory for barcode images
    temp_dir = tempfile.mkdtemp(prefix="barcode_pdf_")
    
    try:
        # Generate barcode images
        barcode_images = {}
        print("🖼️  Generating barcode images...")
        
        for number in numbers:
            try:
                barcode_img = generate_barcode_image(number, width=180, height=80)
                barcode_path = os.path.join(temp_dir, f"barcode_{number}.png")
                barcode_img.save(barcode_path, 'PNG')
                barcode_images[number] = barcode_path
            except Exception as e:
                print(f"⚠️  Could not generate image for {number}: {e}")
        
        print(f"✅ Generated {len(barcode_images)} barcode images")
        
        # Create PDF with A4 size (210mm x 297mm)
        pdf_path = pdf_file
        c = canvas.Canvas(pdf_path, pagesize=A4)
        page_width, page_height = A4
        
        # Calculate spacing for 4 columns x 8 rows layout
        margin_x = 12 * mm  # Left margin
        margin_y = 15 * mm  # Top margin
        
        # Available space
        available_width = page_width - (2 * margin_x)
        available_height = page_height - (2 * margin_y)
        
        # Cell size
        cell_width = available_width / cols
        cell_height = available_height / rows
        
        print(f"\n📐 PDF Layout:")
        print(f"   Columns: {cols}, Rows: {rows}")
        print(f"   Cell size: {cell_width/mm:.1f}mm x {cell_height/mm:.1f}mm")
        
        # Add barcodes to PDF
        barcode_count = 0
        page_num = 1
        
        for idx, number in enumerate(numbers):
            if number not in barcode_images:
                continue
            
            # Calculate position on page
            row = (idx % (cols * rows)) // cols
            col = (idx % (cols * rows)) % cols
            
            # If we've filled this page, create a new one
            if idx > 0 and idx % (cols * rows) == 0:
                c.showPage()
                page_num += 1
                row = 0
                col = 0
            
            # Calculate pixel position
            x = margin_x + (col * cell_width) + (cell_width * 0.05)
            y = page_height - margin_y - ((row + 1) * cell_height) + (cell_height * 0.05)
            
            # Width and height for barcode in PDF
            barcode_width = cell_width * 0.9
            barcode_height = cell_height * 0.75
            
            # Draw barcode image
            barcode_path = barcode_images[number]
            try:
                c.drawImage(barcode_path, x, y, width=barcode_width, height=barcode_height, preserveAspectRatio=True)
                
                # Draw number below barcode
                text_y = y - (cell_height * 0.15)
                c.setFont("Helvetica", 8)
                c.drawString(x + (barcode_width / 2) - 15, text_y, number)
                
                barcode_count += 1
                
            except Exception as e:
                print(f"⚠️  Could not add {number} to PDF: {e}")
        
        # Save PDF
        c.save()
        print(f"\n✅ PDF created: {pdf_path}")
        print(f"📊 Total barcodes in PDF: {barcode_count}")
        print(f"📄 Total pages: {page_num}")
        
    finally:
        # Clean up temporary files
        try:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
                print(f"🧹 Cleaned up temporary files")
        except Exception as e:
            print(f"⚠️  Could not clean temp folder: {e}")


def generate_barcodes_from_sheet(input_file="barcodes_with_images.xlsx", output_file=None):
    """Read numbers from existing Excel sheet and generate barcodes for any missing ones"""
    
    if output_file is None:
        output_file = input_file
    
    # Check if file exists
    if not os.path.exists(input_file):
        print(f"❌ File not found: {input_file}")
        print("Creating new file with initial numbers...")
        numbers = [253320001, 253320002, 253320003, 253320004, 253320005,
                   253320006, 253320007, 253320008, 253320009, 253320010]
        generate_barcodes_excel(numbers, output_file)
        return
    
    # Load existing workbook
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    
    # Create temporary directory for barcode images
    temp_dir = tempfile.mkdtemp(prefix="barcode_")
    print(f"📁 Using temporary directory: {temp_dir}")
    
    try:
        # Find all numbers in column A (starting from row 2)
        existing_numbers = set()
        max_row = ws.max_row
        
        for row in range(2, max_row + 1):
            cell_value = ws[f'A{row}'].value
            if cell_value is not None:
                try:
                    existing_numbers.add(int(cell_value))
                except (ValueError, TypeError):
                    pass
        
        print(f"📊 Found {len(existing_numbers)} existing numbers")
        
        # Check which numbers need barcodes (where column B is empty)
        numbers_needing_barcodes = []
        
        for row in range(2, max_row + 1):
            number_cell = ws[f'A{row}'].value
            barcode_cell = ws[f'B{row}'].value
            
            if number_cell is not None and barcode_cell is None:
                try:
                    number = int(number_cell)
                    numbers_needing_barcodes.append((row, number))
                except (ValueError, TypeError):
                    pass
        
        if not numbers_needing_barcodes:
            print("✅ All numbers already have barcodes!")
            return
        
        print(f"🔄 Generating barcodes for {len(numbers_needing_barcodes)} new numbers...")
        print("=" * 50)
        
        # Generate barcodes for numbers that don't have them
        for row, number in numbers_needing_barcodes:
            try:
                # Generate barcode image
                barcode_img = generate_barcode_image(str(number), width=250, height=100)
                
                # Save to temporary file
                barcode_filename = os.path.join(temp_dir, f"barcode_{number}.png")
                barcode_img.save(barcode_filename, 'PNG')
                
                # Verify file was created
                if os.path.exists(barcode_filename) and os.path.getsize(barcode_filename) > 0:
                    # Insert image into column B
                    img = XLImage(barcode_filename)
                    img.width = 250
                    img.height = 80
                    ws.add_image(img, f'B{row}')
                    
                    # Add barcode representation in column C
                    ws[f'C{row}'] = f"Code128: {number}"
                    ws[f'C{row}'].alignment = Alignment(horizontal="left", vertical="center")
                    
                    # Set row height
                    ws.row_dimensions[row].height = 90
                    
                    print(f"✅ Generated barcode for {number} (Row {row})")
                else:
                    raise Exception("Barcode image file was not created or is empty")
                    
            except Exception as e:
                print(f"❌ Error with {number}: {e}")
                ws[f'B{row}'] = f"Error: {str(e)[:50]}"
                ws[f'C{row}'] = f"Error: {str(e)[:50]}"
        
        # Save workbook
        wb.save(output_file)
        print(f"\n✅ Excel file updated: {output_file}")
        print(f"📊 Total barcodes generated: {len(numbers_needing_barcodes)}")
        
    finally:
        # Clean up temporary files
        try:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
                print(f"🧹 Cleaned up temporary files")
        except Exception as e:
            print(f"⚠️  Could not clean temp folder: {e}")


def generate_barcodes_excel(numbers, output_filename="barcodes_with_images.xlsx"):
    """Generate barcodes for a list of numbers and create Excel file"""
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Barcodes"
    
    # Set headers with formatting
    ws['A1'] = "Number"
    ws['B1'] = "Barcode"
    ws['C1'] = "Barcode Code"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for col in ['A1', 'B1', 'C1']:
        ws[col].fill = header_fill
        ws[col].font = header_font
        ws[col].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.row_dimensions[1].height = 25
    
    # Create temporary directory for saving barcode images
    temp_dir = tempfile.mkdtemp(prefix="barcode_")
    print(f"📁 Using temporary directory: {temp_dir}")
    
    try:
        for idx, number in enumerate(numbers, start=2):
            # Add number to column A
            ws[f'A{idx}'] = number
            
            try:
                # Generate barcode image
                barcode_img = generate_barcode_image(str(number), width=250, height=100)
                
                # Save to temporary file
                barcode_filename = os.path.join(temp_dir, f"barcode_{number}.png")
                barcode_img.save(barcode_filename, 'PNG')
                
                # Verify file was created
                if os.path.exists(barcode_filename) and os.path.getsize(barcode_filename) > 0:
                    # Insert image into column B
                    img = XLImage(barcode_filename)
                    img.width = 250
                    img.height = 80
                    ws.add_image(img, f'B{idx}')
                    
                    # Add barcode representation in column C
                    ws[f'C{idx}'] = f"Code128: {number}"
                    ws[f'C{idx}'].alignment = Alignment(horizontal="left", vertical="center")
                    
                    # Set row height
                    ws.row_dimensions[idx].height = 90
                    
                    print(f"✅ Generated barcode for {number}")
                else:
                    raise Exception("Barcode image file was not created or is empty")
                    
            except Exception as e:
                print(f"❌ Error with {number}: {e}")
                ws[f'B{idx}'] = f"Error: {str(e)[:50]}"
                ws[f'C{idx}'] = f"Error: {str(e)[:50]}"
        
        # Save workbook
        wb.save(output_filename)
        print(f"\n✅ Excel file created: {output_filename}")
        print(f"📊 Total barcodes generated: {len(numbers)}")
        
    finally:
        # Clean up temporary files
        try:
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir)
                print(f"🧹 Cleaned up temporary files")
        except Exception as e:
            print(f"⚠️  Could not clean temp folder: {e}")


if __name__ == "__main__":
    # Check if user wants to create new file or update existing
    print("🔄 Barcode Generator for Excel & PDF")
    print("=" * 50)
    
    choice = input("\nOptions:\n1. Create new Excel file with initial numbers\n2. Update existing Excel file\n3. Generate A4 PDF from existing Excel (4x8 layout)\n\nEnter choice (1, 2, or 3): ").strip()
    
    if choice == "3":
        # PDF generation mode
        excel_file = input("Enter Excel filename (default: barcodes_with_images.xlsx): ").strip() or "barcodes_with_images.xlsx"
        pdf_file = input("Enter PDF filename (default: barcodes_a4.pdf): ").strip() or "barcodes_a4.pdf"
        
        cols = input("Enter number of columns (default: 4): ").strip() or "4"
        rows = input("Enter number of rows (default: 8): ").strip() or "8"
        
        try:
            cols = int(cols)
            rows = int(rows)
        except ValueError:
            cols, rows = 4, 8
        
        generate_pdf_from_excel(excel_file, pdf_file, cols, rows)
        
    elif choice == "2":
        # Update mode - reads existing file and generates barcodes for new numbers
        input_file = input("Enter existing Excel filename (default: barcodes_with_images.xlsx): ").strip() or "barcodes_with_images.xlsx"
        output_file = input("Enter output filename (press Enter to overwrite): ").strip()
        
        if not output_file:
            output_file = input_file
        
        generate_barcodes_from_sheet(input_file, output_file)
        
    else:
        # Create mode - generates initial file
        print("\n📝 Create New Excel File")
        print("=" * 50)
        
        # Ask for output filename
        output_filename = input("Enter Excel filename (default: barcodes_with_images.xlsx): ").strip() or "barcodes_with_images.xlsx"
        
        # Ask for starting number
        start_num = input("Enter starting number (default: 253310001): ").strip() or "253310001"
        try:
            start_num = int(start_num)
        except ValueError:
            start_num = 253310001
            print(f"⚠️  Invalid number, using default: {start_num}")
        
        # Ask for count
        count = input("Enter count of numbers (default: 10): ").strip() or "10"
        try:
            count = int(count)
            if count < 1 or count > 1000:
                raise ValueError("Count must be between 1 and 1000")
        except ValueError:
            count = 10
            print(f"⚠️  Invalid count, using default: {count}")
        
        numbers = [start_num + i for i in range(count)]
        
        print(f"\n🔄 Converting {count} numbers to barcodes...")
        print(f"📊 Range: {start_num} to {start_num + count - 1}")
        print("=" * 50)
        
        generate_barcodes_excel(numbers, output_filename)
    
    print("\n📝 Usage Instructions:")
    print("1. Install required packages:")
    print("   pip install openpyxl pillow python-barcode reportlab")
    print("\n2. Run the script:")
    print("   python generate_barcodes.py")
    print("\n3. Choose an option:")
    print("   Option 1: Create new Excel file")
    print("   Option 2: Update existing Excel file")
    print("   Option 3: Generate A4 PDF (4x8 layout)")
    print("\n✨ Features:")
    print("   • Custom Excel filename")
    print("   • Custom starting number")
    print("   • Custom count (1-1000)")
    print("   • Auto-generate barcodes")
    print("   • Create A4 PDF with 4x8 layout")
    print("   • 32 barcodes per A4 page")
    print("   • Ready to print!")